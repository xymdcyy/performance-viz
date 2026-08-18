import json
from openpyxl import load_workbook

wb = load_workbook(r"d:\代码\业绩分析可视化_26年2月\26年2月KA及智屏渠道业绩分析表格.xlsx", data_only=False)
ws = wb["Sheet8"]

# 1. 基本信息
print(f"Sheet: {ws.title}")
print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")

# 2. 合并单元格
print("\n=== Merged Cells ===")
for mr in ws.merged_cells.ranges:
    tl = mr.start_cell
    val = ws.cell(tl.row, tl.column).value
    print(f"  {mr}: value={repr(val)}")

# 3. 前 15 行内容预览
print("\n=== First 15 Rows ===")
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=False), 1):
    cells = []
    for cell in row:
        v = cell.value
        if v is not None and isinstance(v, str) and v.startswith("="):
            cells.append(f"[FORMULA]{v}")
        else:
            cells.append(repr(v))
    print(f"  Row {row_idx}: {cells}")

# 4. 公式单元格
print("\n=== Formulas ===")
formula_count = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
            formula_count += 1
            if formula_count <= 10:
                print(f"  {cell.coordinate}: {cell.value}")
if formula_count > 10:
    print(f"  ... (共 {formula_count} 个公式)")

# 5. 数据区域检测
print("\n=== Header Detection ===")
for row_idx in range(1, min(15, ws.max_row + 1)):
    row_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
    non_none = [v for v in row_values if v is not None]
    types = [type(v).__name__ for v in non_none]
    print(f"  Row {row_idx}: {len(non_none)} non-empty, types={set(types)}")

# 6. 列宽信息
print("\n=== Column Info ===")
for col_idx in range(1, min(20, ws.max_column + 1)):
    col_letter = ws.cell(1, col_idx).column_letter
    val = ws.cell(1, col_idx).value
    print(f"  Col {col_letter} ({col_idx}): {repr(val)}")
