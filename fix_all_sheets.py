import pandas as pd
import json
import os
import warnings
import re
import openpyxl
warnings.filterwarnings('ignore')

EXCEL_FILE = 'd:/代码/26年2月KA及智屏渠道业绩分析表格.xlsx'
OUTPUT_DIR = 'd:/代码/业绩分析可视化_26年2月'

if not os.path.exists(os.path.join(OUTPUT_DIR, 'json')):
    os.makedirs(os.path.join(OUTPUT_DIR, 'json'))

def save_json(data, filename):
    filepath = os.path.join(OUTPUT_DIR, 'json', filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已保存: {filepath}")

def safe_float(value):
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value)
    match = re.search(r'-?\d+\.?\d*', value_str)
    if match:
        try:
            return float(match.group())
        except:
            return None
    return None

def extract_change_value(value):
    if pd.isna(value):
        return None
    value_str = str(value)
    match = re.search(r'-?\d+\.?\d*', value_str)
    if match:
        val = float(match.group())
        if '减少' in value_str or '增亏' in value_str:
            return -val
        if '增加' in value_str or '减亏' in value_str:
            return val
        return val
    return None

def get_cell_value(sheet, row, col):
    try:
        val = sheet.cell(row=row, column=col).value
        return val
    except:
        return None

def process_sheet1():
    print("处理Sheet1...")
    wb_xl = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb_xl['Sheet1']
    
    data = {
        "sheet_name": "Sheet1",
        "title": "KA及智屏渠道业绩分析（26年2月）",
        "data": []
    }
    
    row_mapping = [
        ('KA合计', '合计', 3),
        ('KA合计', '全国连锁', 4),
        ('KA合计', '区域连锁', 5),
        ('智屏-TCL', '合计', 6),
        ('智屏-TCL', '全国连锁', 7),
        ('智屏-TCL', '区域连锁', 8),
        ('空调-TCL', '合计', 9),
        ('空调-TCL', '全国连锁', 10),
        ('空调-TCL', '区域连锁', 11),
        ('冰洗-TCL', '合计', 12),
        ('冰洗-TCL', '全国连锁', 13),
        ('冰洗-TCL', '区域连锁', 14),
        ('CIOT-TCL', '合计', 15),
        ('CIOT-TCL', '全国连锁', 16),
        ('CIOT-TCL', '区域连锁', 17),
    ]
    
    for category, channel, row_idx in row_mapping:
        record = {
            "品类": category,
            "渠道": channel,
            "收入_实际": safe_float(get_cell_value(sheet, row_idx, 2)),
            "收入_BP": safe_float(get_cell_value(sheet, row_idx, 3)),
            "收入_BP达成率": safe_float(get_cell_value(sheet, row_idx, 4)),
            "收入_同期": safe_float(get_cell_value(sheet, row_idx, 5)),
            "收入_同比": safe_float(get_cell_value(sheet, row_idx, 6)),
            "销量_实际": safe_float(get_cell_value(sheet, row_idx, 7)),
            "销量_BP达成率": safe_float(get_cell_value(sheet, row_idx, 8)),
            "销量_同比": safe_float(get_cell_value(sheet, row_idx, 9)),
            "ASP_实际": safe_float(get_cell_value(sheet, row_idx, 10)),
            "ASP_BP达成": safe_float(get_cell_value(sheet, row_idx, 11)),
            "ASP_同比": safe_float(get_cell_value(sheet, row_idx, 12)),
            "端到端毛利率_实际": safe_float(get_cell_value(sheet, row_idx, 13)),
            "端到端毛利率_BP达成率": extract_change_value(get_cell_value(sheet, row_idx, 14)),
            "端到端毛利率_同比": extract_change_value(get_cell_value(sheet, row_idx, 15)),
            "端到端净利_实际": safe_float(get_cell_value(sheet, row_idx, 16)),
            "端到端净利_BP达成率": extract_change_value(get_cell_value(sheet, row_idx, 17)),
            "端到端净利_同比": extract_change_value(get_cell_value(sheet, row_idx, 18)),
        }
        data["data"].append(record)
    
    save_json(data, 'sheet1.json')
    return data

def process_sheet2():
    print("处理Sheet2...")
    wb_xl = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb_xl['Sheet2']
    
    data = {
        "sheet_name": "Sheet2",
        "title": "KA及智屏渠道业绩分析（1-2月累计）",
        "data": []
    }
    
    row_mapping = [
        ('KA合计', '合计', 3),
        ('KA合计', '全国连锁', 4),
        ('KA合计', '区域连锁', 5),
        ('智屏-TCL', '合计', 6),
        ('智屏-TCL', '全国连锁', 7),
        ('智屏-TCL', '区域连锁', 8),
        ('空调-TCL', '合计', 9),
        ('空调-TCL', '全国连锁', 10),
        ('空调-TCL', '区域连锁', 11),
        ('冰洗-TCL', '合计', 12),
        ('冰洗-TCL', '全国连锁', 13),
        ('冰洗-TCL', '区域连锁', 14),
        ('CIOT-TCL', '合计', 15),
        ('CIOT-TCL', '全国连锁', 16),
        ('CIOT-TCL', '区域连锁', 17),
    ]
    
    for category, channel, row_idx in row_mapping:
        record = {
            "品类": category,
            "渠道": channel,
            "收入_实际": safe_float(get_cell_value(sheet, row_idx, 2)),
            "收入_BP": safe_float(get_cell_value(sheet, row_idx, 3)),
            "收入_BP达成率": safe_float(get_cell_value(sheet, row_idx, 4)),
            "收入_同期": safe_float(get_cell_value(sheet, row_idx, 5)),
            "收入_同比": safe_float(get_cell_value(sheet, row_idx, 6)),
            "销量_实际": safe_float(get_cell_value(sheet, row_idx, 7)),
            "销量_BP达成率": safe_float(get_cell_value(sheet, row_idx, 8)),
            "销量_同比": safe_float(get_cell_value(sheet, row_idx, 9)),
            "ASP_实际": safe_float(get_cell_value(sheet, row_idx, 10)),
            "ASP_BP达成": safe_float(get_cell_value(sheet, row_idx, 11)),
            "ASP_同比": safe_float(get_cell_value(sheet, row_idx, 12)),
            "端到端毛利率_实际": safe_float(get_cell_value(sheet, row_idx, 13)),
            "端到端毛利率_BP达成率": extract_change_value(get_cell_value(sheet, row_idx, 14)),
            "端到端毛利率_同比": extract_change_value(get_cell_value(sheet, row_idx, 15)),
            "端到端净利_实际": safe_float(get_cell_value(sheet, row_idx, 16)),
            "端到端净利_BP达成率": extract_change_value(get_cell_value(sheet, row_idx, 17)),
            "端到端净利_同比": extract_change_value(get_cell_value(sheet, row_idx, 18)),
        }
        data["data"].append(record)
    
    save_json(data, 'sheet2.json')
    return data

def process_sheet3():
    print("处理Sheet3...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Sheet3']
    
    data = {
        "sheet_name": "Sheet3",
        "title": "费用项目分析",
        "data": []
    }
    
    # 收入数据: Row4
    record = {
        "费用项目": "收入",
        "类型": "金额",
        "合计_实际": safe_float(get_cell_value(sheet, 4, 3)),
        "智屏_连锁渠道_实际": safe_float(get_cell_value(sheet, 4, 4)),
        "智屏_区域连锁_实际": safe_float(get_cell_value(sheet, 4, 5)),
        "空调_连锁渠道_实际": safe_float(get_cell_value(sheet, 4, 6)),
        "空调_区域连锁_实际": safe_float(get_cell_value(sheet, 4, 7)),
        "冰洗_连锁渠道_实际": safe_float(get_cell_value(sheet, 4, 8)),
        "冰洗_区域连锁_实际": safe_float(get_cell_value(sheet, 4, 9)),
        "CIOT_连锁渠道_实际": safe_float(get_cell_value(sheet, 4, 10)),
        "CIOT_区域连锁_实际": safe_float(get_cell_value(sheet, 4, 11)),
        "合计_BP比": safe_float(get_cell_value(sheet, 4, 12)),
        "智屏_连锁渠道_BP比": safe_float(get_cell_value(sheet, 4, 13)),
        "智屏_区域连锁_BP比": safe_float(get_cell_value(sheet, 4, 14)),
        "空调_连锁渠道_BP比": safe_float(get_cell_value(sheet, 4, 15)),
        "空调_区域连锁_BP比": safe_float(get_cell_value(sheet, 4, 16)),
        "冰洗_连锁渠道_BP比": safe_float(get_cell_value(sheet, 4, 17)),
        "冰洗_区域连锁_BP比": safe_float(get_cell_value(sheet, 4, 18)),
        "CIOT_连锁渠道_BP比": safe_float(get_cell_value(sheet, 4, 19)),
        "CIOT_区域连锁_BP比": safe_float(get_cell_value(sheet, 4, 20)),
        "合计_同比": safe_float(get_cell_value(sheet, 4, 21)),
        "智屏_连锁渠道_同比": safe_float(get_cell_value(sheet, 4, 22)),
        "智屏_区域连锁_同比": safe_float(get_cell_value(sheet, 4, 23)),
        "空调_连锁渠道_同比": safe_float(get_cell_value(sheet, 4, 24)),
        "空调_区域连锁_同比": safe_float(get_cell_value(sheet, 4, 25)),
        "冰洗_连锁渠道_同比": safe_float(get_cell_value(sheet, 4, 26)),
        "冰洗_区域连锁_同比": safe_float(get_cell_value(sheet, 4, 27)),
        "CIOT_连锁渠道_同比": safe_float(get_cell_value(sheet, 4, 28)),
        "CIOT_区域连锁_同比": safe_float(get_cell_value(sheet, 4, 29)),
    }
    data["data"].append(record)
    
    # 金额数据: Row5-12 (8项)
    expense_items = {
        5: '汇总费用',
        6: '固定性费用',
        7: '人力成本',
        8: '日常费用',
        9: '研发费用',
        10: '财务费用',
        11: '管理费用',
        12: '变动性费用',
    }
    
    for row_idx, item_name in expense_items.items():
        record = {
            "费用项目": item_name,
            "类型": "金额",
            "合计_实际": safe_float(get_cell_value(sheet, row_idx, 3)),
            "智屏_连锁渠道_实际": safe_float(get_cell_value(sheet, row_idx, 4)),
            "智屏_区域连锁_实际": safe_float(get_cell_value(sheet, row_idx, 5)),
            "空调_连锁渠道_实际": safe_float(get_cell_value(sheet, row_idx, 6)),
            "空调_区域连锁_实际": safe_float(get_cell_value(sheet, row_idx, 7)),
            "冰洗_连锁渠道_实际": safe_float(get_cell_value(sheet, row_idx, 8)),
            "冰洗_区域连锁_实际": safe_float(get_cell_value(sheet, row_idx, 9)),
            "CIOT_连锁渠道_实际": safe_float(get_cell_value(sheet, row_idx, 10)),
            "CIOT_区域连锁_实际": safe_float(get_cell_value(sheet, row_idx, 11)),
            "合计_BP比": safe_float(get_cell_value(sheet, row_idx, 12)),
            "智屏_连锁渠道_BP比": safe_float(get_cell_value(sheet, row_idx, 13)),
            "智屏_区域连锁_BP比": safe_float(get_cell_value(sheet, row_idx, 14)),
            "空调_连锁渠道_BP比": safe_float(get_cell_value(sheet, row_idx, 15)),
            "空调_区域连锁_BP比": safe_float(get_cell_value(sheet, row_idx, 16)),
            "冰洗_连锁渠道_BP比": safe_float(get_cell_value(sheet, row_idx, 17)),
            "冰洗_区域连锁_BP比": safe_float(get_cell_value(sheet, row_idx, 18)),
            "CIOT_连锁渠道_BP比": safe_float(get_cell_value(sheet, row_idx, 19)),
            "CIOT_区域连锁_BP比": safe_float(get_cell_value(sheet, row_idx, 20)),
            "合计_同比": safe_float(get_cell_value(sheet, row_idx, 21)),
            "智屏_连锁渠道_同比": safe_float(get_cell_value(sheet, row_idx, 22)),
            "智屏_区域连锁_同比": safe_float(get_cell_value(sheet, row_idx, 23)),
            "空调_连锁渠道_同比": safe_float(get_cell_value(sheet, row_idx, 24)),
            "空调_区域连锁_同比": safe_float(get_cell_value(sheet, row_idx, 25)),
            "冰洗_连锁渠道_同比": safe_float(get_cell_value(sheet, row_idx, 26)),
            "冰洗_区域连锁_同比": safe_float(get_cell_value(sheet, row_idx, 27)),
            "CIOT_连锁渠道_同比": safe_float(get_cell_value(sheet, row_idx, 28)),
            "CIOT_区域连锁_同比": safe_float(get_cell_value(sheet, row_idx, 29)),
        }
        data["data"].append(record)
    
    # 费率数据: Row13-21 (9项)
    rate_items = {
        13: '汇总费用费率',
        14: '固定性费用费率',
        15: '变动性费用费率',
        16: '激励及佣金费率',
        17: '储运费费率',
        18: '市场推广费率',
        19: '售后服务费率',
        20: '安装费费率',
        21: '品牌费用费率',
    }
    
    for row_idx, item_name in rate_items.items():
        record = {
            "费用项目": item_name,
            "类型": "费率",
            "合计_实际": safe_float(get_cell_value(sheet, row_idx, 3)),
            "智屏_连锁渠道_实际": safe_float(get_cell_value(sheet, row_idx, 4)),
            "智屏_区域连锁_实际": safe_float(get_cell_value(sheet, row_idx, 5)),
            "空调_连锁渠道_实际": safe_float(get_cell_value(sheet, row_idx, 6)),
            "空调_区域连锁_实际": safe_float(get_cell_value(sheet, row_idx, 7)),
            "冰洗_连锁渠道_实际": safe_float(get_cell_value(sheet, row_idx, 8)),
            "冰洗_区域连锁_实际": safe_float(get_cell_value(sheet, row_idx, 9)),
            "CIOT_连锁渠道_实际": safe_float(get_cell_value(sheet, row_idx, 10)),
            "CIOT_区域连锁_实际": safe_float(get_cell_value(sheet, row_idx, 11)),
            "合计_BP比": safe_float(get_cell_value(sheet, row_idx, 12)),
            "智屏_连锁渠道_BP比": safe_float(get_cell_value(sheet, row_idx, 13)),
            "智屏_区域连锁_BP比": safe_float(get_cell_value(sheet, row_idx, 14)),
            "空调_连锁渠道_BP比": safe_float(get_cell_value(sheet, row_idx, 15)),
            "空调_区域连锁_BP比": safe_float(get_cell_value(sheet, row_idx, 16)),
            "冰洗_连锁渠道_BP比": safe_float(get_cell_value(sheet, row_idx, 17)),
            "冰洗_区域连锁_BP比": safe_float(get_cell_value(sheet, row_idx, 18)),
            "CIOT_连锁渠道_BP比": safe_float(get_cell_value(sheet, row_idx, 19)),
            "CIOT_区域连锁_BP比": safe_float(get_cell_value(sheet, row_idx, 20)),
            "合计_同比": safe_float(get_cell_value(sheet, row_idx, 21)),
            "智屏_连锁渠道_同比": safe_float(get_cell_value(sheet, row_idx, 22)),
            "智屏_区域连锁_同比": safe_float(get_cell_value(sheet, row_idx, 23)),
            "空调_连锁渠道_同比": safe_float(get_cell_value(sheet, row_idx, 24)),
            "空调_区域连锁_同比": safe_float(get_cell_value(sheet, row_idx, 25)),
            "冰洗_连锁渠道_同比": safe_float(get_cell_value(sheet, row_idx, 26)),
            "冰洗_区域连锁_同比": safe_float(get_cell_value(sheet, row_idx, 27)),
            "CIOT_连锁渠道_同比": safe_float(get_cell_value(sheet, row_idx, 28)),
            "CIOT_区域连锁_同比": safe_float(get_cell_value(sheet, row_idx, 29)),
        }
        data["data"].append(record)
    
    save_json(data, 'sheet3.json')
    return data

def process_sheet4():
    print("处理Sheet4...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Sheet4']
    
    data = {
        "sheet_name": "Sheet4",
        "title": "产品线费用分解",
        "data": []
    }
    
    # 结构: 产品名+渠道, 当期(Row), 同期(Row+1), 同期比(Row+2)
    products = [
        (3, '合计', '汇总'),
        (6, '合计', '连锁'),
        (9, '合计', '区域连锁'),
        (12, '智屏', '汇总'),
        (15, '智屏', '连锁'),
        (18, '智屏', '区域连锁'),
        (21, '空调', '汇总'),
        (24, '空调', '连锁'),
        (27, '空调', '区域连锁'),
        (30, '白电', '汇总'),
        (33, '白电', '连锁'),
        (36, '白电', '区域连锁'),
    ]
    
    for start_row, product, channel in products:
        project = get_cell_value(sheet, start_row, 3)  # 当期/同期/同期比
        
        # 当期
        record = {
            "产品线": product,
            "渠道": channel,
            "项目": "当期",
            "毛收入": safe_float(get_cell_value(sheet, start_row, 4)),
            "固定费用": safe_float(get_cell_value(sheet, start_row, 5)),
            "固定费用_率": safe_float(get_cell_value(sheet, start_row, 6)),
            "变动费用": safe_float(get_cell_value(sheet, start_row, 7)),
            "变动费用_率": safe_float(get_cell_value(sheet, start_row, 8)),
            "可控费用": safe_float(get_cell_value(sheet, start_row, 9)),
            "可控费用_率": safe_float(get_cell_value(sheet, start_row, 10)),
            "客户费用合计": safe_float(get_cell_value(sheet, start_row, 11)),
            "客户费用合计_率": safe_float(get_cell_value(sheet, start_row, 12)),
            "折扣": safe_float(get_cell_value(sheet, start_row, 13)),
            "折扣_率": safe_float(get_cell_value(sheet, start_row, 14)),
            "财务支持": safe_float(get_cell_value(sheet, start_row, 15)),
            "财务支持_率": safe_float(get_cell_value(sheet, start_row, 16)),
            "定价空间": safe_float(get_cell_value(sheet, start_row, 17)),
            "定价空间_率": safe_float(get_cell_value(sheet, start_row, 18)),
            "价格保护": safe_float(get_cell_value(sheet, start_row, 19)),
            "价格保护_率": safe_float(get_cell_value(sheet, start_row, 20)),
            "客户支持": safe_float(get_cell_value(sheet, start_row, 21)),
            "客户支持_率": safe_float(get_cell_value(sheet, start_row, 22)),
            "联合促销与推广": safe_float(get_cell_value(sheet, start_row, 23)),
            "联合促销与推广_率": safe_float(get_cell_value(sheet, start_row, 24)),
            "临时激励": safe_float(get_cell_value(sheet, start_row, 25)),
            "临时激励_率": safe_float(get_cell_value(sheet, start_row, 26)),
            "销售返利": safe_float(get_cell_value(sheet, start_row, 27)),
            "销售返利_率": safe_float(get_cell_value(sheet, start_row, 28)),
        }
        data["data"].append(record)
        
        # 同期
        record2 = {
            "产品线": product,
            "渠道": channel,
            "项目": "同期",
            "毛收入": safe_float(get_cell_value(sheet, start_row + 1, 4)),
            "固定费用": safe_float(get_cell_value(sheet, start_row + 1, 5)),
            "固定费用_率": safe_float(get_cell_value(sheet, start_row + 1, 6)),
            "变动费用": safe_float(get_cell_value(sheet, start_row + 1, 7)),
            "变动费用_率": safe_float(get_cell_value(sheet, start_row + 1, 8)),
            "可控费用": safe_float(get_cell_value(sheet, start_row + 1, 9)),
            "可控费用_率": safe_float(get_cell_value(sheet, start_row + 1, 10)),
            "客户费用合计": safe_float(get_cell_value(sheet, start_row + 1, 11)),
            "客户费用合计_率": safe_float(get_cell_value(sheet, start_row + 1, 12)),
            "折扣": safe_float(get_cell_value(sheet, start_row + 1, 13)),
            "折扣_率": safe_float(get_cell_value(sheet, start_row + 1, 14)),
            "财务支持": safe_float(get_cell_value(sheet, start_row + 1, 15)),
            "财务支持_率": safe_float(get_cell_value(sheet, start_row + 1, 16)),
            "定价空间": safe_float(get_cell_value(sheet, start_row + 1, 17)),
            "定价空间_率": safe_float(get_cell_value(sheet, start_row + 1, 18)),
            "价格保护": safe_float(get_cell_value(sheet, start_row + 1, 19)),
            "价格保护_率": safe_float(get_cell_value(sheet, start_row + 1, 20)),
            "客户支持": safe_float(get_cell_value(sheet, start_row + 1, 21)),
            "客户支持_率": safe_float(get_cell_value(sheet, start_row + 1, 22)),
            "联合促销与推广": safe_float(get_cell_value(sheet, start_row + 1, 23)),
            "联合促销与推广_率": safe_float(get_cell_value(sheet, start_row + 1, 24)),
            "临时激励": safe_float(get_cell_value(sheet, start_row + 1, 25)),
            "临时激励_率": safe_float(get_cell_value(sheet, start_row + 1, 26)),
            "销售返利": safe_float(get_cell_value(sheet, start_row + 1, 27)),
            "销售返利_率": safe_float(get_cell_value(sheet, start_row + 1, 28)),
        }
        data["data"].append(record2)
        
        # 同期比
        record3 = {
            "产品线": product,
            "渠道": channel,
            "项目": "同期比",
            "毛收入": safe_float(get_cell_value(sheet, start_row + 2, 4)),
            "固定费用": safe_float(get_cell_value(sheet, start_row + 2, 5)),
            "固定费用_率": safe_float(get_cell_value(sheet, start_row + 2, 6)),
            "变动费用": safe_float(get_cell_value(sheet, start_row + 2, 7)),
            "变动费用_率": safe_float(get_cell_value(sheet, start_row + 2, 8)),
            "可控费用": safe_float(get_cell_value(sheet, start_row + 2, 9)),
            "可控费用_率": safe_float(get_cell_value(sheet, start_row + 2, 10)),
            "客户费用合计": safe_float(get_cell_value(sheet, start_row + 2, 11)),
            "客户费用合计_率": safe_float(get_cell_value(sheet, start_row + 2, 12)),
            "折扣": safe_float(get_cell_value(sheet, start_row + 2, 13)),
            "折扣_率": safe_float(get_cell_value(sheet, start_row + 2, 14)),
            "财务支持": safe_float(get_cell_value(sheet, start_row + 2, 15)),
            "财务支持_率": safe_float(get_cell_value(sheet, start_row + 2, 16)),
            "定价空间": safe_float(get_cell_value(sheet, start_row + 2, 17)),
            "定价空间_率": safe_float(get_cell_value(sheet, start_row + 2, 18)),
            "价格保护": safe_float(get_cell_value(sheet, start_row + 2, 19)),
            "价格保护_率": safe_float(get_cell_value(sheet, start_row + 2, 20)),
            "客户支持": safe_float(get_cell_value(sheet, start_row + 2, 21)),
            "客户支持_率": safe_float(get_cell_value(sheet, start_row + 2, 22)),
            "联合促销与推广": safe_float(get_cell_value(sheet, start_row + 2, 23)),
            "联合促销与推广_率": safe_float(get_cell_value(sheet, start_row + 2, 24)),
            "临时激励": safe_float(get_cell_value(sheet, start_row + 2, 25)),
            "临时激励_率": safe_float(get_cell_value(sheet, start_row + 2, 26)),
            "销售返利": safe_float(get_cell_value(sheet, start_row + 2, 27)),
            "销售返利_率": safe_float(get_cell_value(sheet, start_row + 2, 28)),
        }
        data["data"].append(record3)
     
    save_json(data, 'sheet4.json')
    return data

def process_sheet5():
    print("处理Sheet5...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Sheet5']
    
    data = {
        "sheet_name": "Sheet5",
        "title": "H1月度预测",
        "data": []
    }
    
    def parse_bp_ratio(val):
        if val is None:
            return None
        s = str(val).strip()
        if not s:
            return None
        try:
            return float(s)
        except:
            return None
    
    def parse_margin_change(val):
        if val is None:
            return None
        s = str(val).strip()
        if not s:
            return None
        if '增加' in s:
            num = s.replace('增加', '').replace('个点', '').strip()
            try:
                return round(float(num) / 100, 4)
            except:
                return None
        elif '减少' in s:
            num = s.replace('减少', '').replace('个点', '').strip()
            try:
                return round(-float(num) / 100, 4)
            except:
                return None
        return None
    
    def parse_profit_change(val):
        if val is None:
            return None
        s = str(val).strip()
        if not s:
            return None
        # 去掉空格
        s = s.replace(' ', '')
        if '减亏' in s:
            num = s.replace('减亏', '').replace('万', '').strip()
            try:
                return float(num)
            except:
                return None
        elif '增亏' in s:
            num = s.replace('增亏', '').replace('万', '').strip()
            try:
                return -float(num)
            except:
                return None
        elif '增盈' in s:
            num = s.replace('增盈', '').replace('万', '').strip()
            try:
                return float(num)
            except:
                return None
        elif '减盈' in s:
            num = s.replace('减盈', '').replace('万', '').strip()
            try:
                return -float(num)
            except:
                return None
        return None
    
    # 结构: 每个品类占3行
    categories = {
        3: '合计',
        6: '智屏',
        9: '空调',
        12: '冰洗',
        15: 'CIOT'
    }
    
    # 数据类型 (行偏移量)
    types = {0: '收入', 1: '毛利率', 2: '利润'}
    
    for cat_row, category in categories.items():
        for type_offset, data_type in types.items():
            row = cat_row + type_offset
            
            # 根据类型选择解析函数
            bp_val = get_cell_value(sheet, row, 4)
            yoy_val = get_cell_value(sheet, row, 5)
            
            if data_type == '毛利率':
                bp_parsed = parse_margin_change(bp_val)
                yoy_parsed = parse_margin_change(yoy_val)
            elif data_type == '利润':
                bp_parsed = parse_profit_change(bp_val)
                yoy_parsed = parse_profit_change(yoy_val)
            else:  # 收入
                bp_parsed = parse_bp_ratio(bp_val)
                yoy_parsed = parse_bp_ratio(yoy_val)
            
            # 其他列也类似处理
            bp3 = get_cell_value(sheet, row, 7)
            yoy3 = get_cell_value(sheet, row, 8)
            bp4 = get_cell_value(sheet, row, 10)
            yoy4 = get_cell_value(sheet, row, 11)
            bp5 = get_cell_value(sheet, row, 13)
            yoy5 = get_cell_value(sheet, row, 14)
            bp6 = get_cell_value(sheet, row, 16)
            yoy6 = get_cell_value(sheet, row, 17)
            bp_h1 = get_cell_value(sheet, row, 19)
            yoy_h1 = get_cell_value(sheet, row, 20)
            
            if data_type == '毛利率':
                bp3 = parse_margin_change(bp3)
                yoy3 = parse_margin_change(yoy3)
                bp4 = parse_margin_change(bp4)
                yoy4 = parse_margin_change(yoy4)
                bp5 = parse_margin_change(bp5)
                yoy5 = parse_margin_change(yoy5)
                bp6 = parse_margin_change(bp6)
                yoy6 = parse_margin_change(yoy6)
                bp_h1 = parse_margin_change(bp_h1)
                yoy_h1 = parse_margin_change(yoy_h1)
            elif data_type == '利润':
                bp3 = parse_profit_change(bp3)
                yoy3 = parse_profit_change(yoy3)
                bp4 = parse_profit_change(bp4)
                yoy4 = parse_profit_change(yoy4)
                bp5 = parse_profit_change(bp5)
                yoy5 = parse_profit_change(yoy5)
                bp6 = parse_profit_change(bp6)
                yoy6 = parse_profit_change(yoy6)
                bp_h1 = parse_profit_change(bp_h1)
                yoy_h1 = parse_profit_change(yoy_h1)
            
            record = {
                "品类": category,
                "类型": data_type,
                "1-2月_实际": safe_float(get_cell_value(sheet, row, 3)),
                "1-2月_BP达成": bp_parsed,
                "1-2月_同比": yoy_parsed,
                "3月_预测": safe_float(get_cell_value(sheet, row, 6)),
                "3月_BP达成": bp3,
                "3月_同比": yoy3,
                "4月_预测": safe_float(get_cell_value(sheet, row, 9)),
                "4月_BP达成": bp4,
                "4月_同比": yoy4,
                "5月_预测": safe_float(get_cell_value(sheet, row, 12)),
                "5月_BP达成": bp5,
                "5月_同比": yoy5,
                "6月_预测": safe_float(get_cell_value(sheet, row, 15)),
                "6月_BP达成": bp6,
                "6月_同比": yoy6,
                "H1累计": safe_float(get_cell_value(sheet, row, 18)),
                "H1_BP达成": bp_h1,
                "H1_同比": yoy_h1,
            }
            data["data"].append(record)
    
    save_json(data, 'sheet5.json')
    return data

def process_sheet6():
    print("处理Sheet6...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Sheet6']
    
    data = {
        "sheet_name": "Sheet6",
        "title": "应收账款逾期分析",
        "data": []
    }
    
    # 先添加合计行
    c1 = get_cell_value(sheet, 4, 1)
    if c1 and '合计' in str(c1):
        data["data"].append({
            "大区": "合计",
            "战区": "",
            "渠道": "",
            "合计_应收": safe_float(get_cell_value(sheet, 4, 4)),
            "合计_超期应收": safe_float(get_cell_value(sheet, 4, 5)),
            "合计_逾期率": safe_float(get_cell_value(sheet, 4, 6)),
            "合计_逾期_环比": safe_float(get_cell_value(sheet, 4, 7)),
            "合计_逾期_同比": safe_float(get_cell_value(sheet, 4, 8)),
            "合计_智屏逾期": safe_float(get_cell_value(sheet, 4, 9)),
            "合计_空调逾期": safe_float(get_cell_value(sheet, 4, 10)),
            "合计_白电逾期": safe_float(get_cell_value(sheet, 4, 11)),
            "合计_CIOT逾期": safe_float(get_cell_value(sheet, 4, 12)),
        })
    
    current_region = ""
    for row_idx in range(5, sheet.max_row + 1):
        c1 = get_cell_value(sheet, row_idx, 1)
        c2 = get_cell_value(sheet, row_idx, 2)
        c3 = get_cell_value(sheet, row_idx, 3)
        
        if c1 is None and c2 is None and c3 is None:
            continue
        
        c1_str = str(c1) if c1 else ""
        c2_str = str(c2) if c2 else ""
        c3_str = str(c3) if c3 else ""
        
        if '合计' in c1_str:
            continue
        if '小计' in c2_str:
            continue
        if '战区' in c1_str:
            continue
        if '其中' in c3_str:
            continue
        
        if c1 and c1 not in ['战区']:
            current_region = c1_str
        
        region = current_region
        warzone = c2_str if c2_str and c2_str not in ['总部', '战区'] else ""
        if c2_str == '总部':
            warzone = "总部"
        
        channel = c3_str
        
        record = {
            "大区": region,
            "战区": warzone,
            "渠道": channel,
            "合计_应收": safe_float(get_cell_value(sheet, row_idx, 4)),
            "合计_超期应收": safe_float(get_cell_value(sheet, row_idx, 5)),
            "合计_逾期率": safe_float(get_cell_value(sheet, row_idx, 6)),
            "逾期趋势_环比": safe_float(get_cell_value(sheet, row_idx, 7)),
            "逾期趋势_同比": safe_float(get_cell_value(sheet, row_idx, 8)),
            "逾期金额_智屏": safe_float(get_cell_value(sheet, row_idx, 9)),
            "逾期金额_空调": safe_float(get_cell_value(sheet, row_idx, 10)),
            "逾期金额_白电": safe_float(get_cell_value(sheet, row_idx, 11)),
            "逾期金额_CIOT": safe_float(get_cell_value(sheet, row_idx, 12)),
        }
        data["data"].append(record)
    
    print(f"Sheet6: 共{len(data['data'])}条记录")
    save_json(data, 'sheet6.json')
    return data

def process_sheet7():
    print("处理Sheet7...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Sheet7']
    
    data = {
        "sheet_name": "Sheet7",
        "title": "客户存销和滞销分析",
        "data": []
    }
    
    current_region = ""
    for row_idx in range(6, sheet.max_row + 1):
        region = get_cell_value(sheet, row_idx, 1)
        warzone = get_cell_value(sheet, row_idx, 2)
        
        if region is None or region == '':
            if warzone and warzone != '小计':
                pass
            else:
                continue
        else:
            if '小计' in str(region) or '合计' in str(region):
                continue
            current_region = str(region)
        
        if warzone and str(warzone) != '小计':
            record = {
                "大区": current_region,
                "战区": str(warzone) if warzone else "",
                "库存_合计": safe_float(get_cell_value(sheet, row_idx, 3)),
                "销售_合计": safe_float(get_cell_value(sheet, row_idx, 4)),
                "存销比_合计": safe_float(get_cell_value(sheet, row_idx, 5)),
                "库存_苏宁": safe_float(get_cell_value(sheet, row_idx, 6)),
                "销售_苏宁": safe_float(get_cell_value(sheet, row_idx, 7)),
                "存销比_苏宁": safe_float(get_cell_value(sheet, row_idx, 8)),
                "库存_五星": safe_float(get_cell_value(sheet, row_idx, 9)),
                "销售_五星": safe_float(get_cell_value(sheet, row_idx, 10)),
                "存销比_五星": safe_float(get_cell_value(sheet, row_idx, 11)),
                "滞销_苏宁_库存_台数": safe_float(get_cell_value(sheet, row_idx, 12)),
                "滞销_苏宁_库存_金额": safe_float(get_cell_value(sheet, row_idx, 13)),
                "滞销_苏宁_60+库存_台数": safe_float(get_cell_value(sheet, row_idx, 14)),
                "滞销_苏宁_60+库存_金额": safe_float(get_cell_value(sheet, row_idx, 15)),
                "滞销_苏宁_滞销占比_台数": safe_float(get_cell_value(sheet, row_idx, 16)),
                "滞销_苏宁_滞销占比_金额": safe_float(get_cell_value(sheet, row_idx, 17)),
                "滞销_五星_库存_台数": safe_float(get_cell_value(sheet, row_idx, 18)),
                "滞销_五星_库存_金额": safe_float(get_cell_value(sheet, row_idx, 19)),
                "滞销_五星_90+库存_台数": safe_float(get_cell_value(sheet, row_idx, 20)),
                "滞销_五星_90+库存_金额": safe_float(get_cell_value(sheet, row_idx, 21)),
                "滞销_五星_滞销占比_台数": safe_float(get_cell_value(sheet, row_idx, 22)),
                "滞销_五星_滞销占比_金额": safe_float(get_cell_value(sheet, row_idx, 23)),
            }
            data["data"].append(record)
    
    save_json(data, 'sheet7.json')
    print(f"Sheet7: 共{len(data['data'])}条记录")
    return data

def process_sheet8():
    print("处理Sheet8...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Sheet8']
    
    data = {
        "sheet_name": "Sheet8",
        "title": "零售vs库存对比",
        "data": []
    }
    
    current_region = ""
    for row_idx in range(3, sheet.max_row + 1):
        region = get_cell_value(sheet, row_idx, 1)
        warzone = get_cell_value(sheet, row_idx, 2)
        
        if region is None or region == '':
            if warzone is None or warzone == '':
                continue
        else:
            if '小计' in str(region) or '合计' in str(region) or region == '总计':
                continue
            current_region = str(region)
        
        if warzone and str(warzone) != '小计':
            record = {
                "大区": current_region,
                "战区": str(warzone) if warzone else "",
                "零售": safe_float(get_cell_value(sheet, row_idx, 3)),
                "库存": safe_float(get_cell_value(sheet, row_idx, 4)),
                "存销比": safe_float(get_cell_value(sheet, row_idx, 5)),
            }
            data["data"].append(record)
    
    save_json(data, 'sheet8.json')
    print(f"Sheet8: 共{len(data['data'])}条记录")
    return data

def process_sheet9():
    print("处理Sheet9...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Sheet9']
    
    data = {
        "sheet_name": "Sheet9",
        "title": "库存金额分布",
        "data": []
    }
    
    current_region = ""
    current_warzone = ""
    for row_idx in range(3, sheet.max_row + 1):
        region = get_cell_value(sheet, row_idx, 1)
        warzone = get_cell_value(sheet, row_idx, 2)
        dept = get_cell_value(sheet, row_idx, 3)
        
        if region is None or region == '':
            pass
        else:
            if '小计' in str(region) or '合计' in str(region):
                continue
            current_region = str(region)
        
        if warzone:
            if '小计' in str(warzone) or '合计' in str(warzone):
                continue
            current_warzone = str(warzone)
        
        if dept is None or dept == '':
            continue
            
        record = {
            "大区": current_region,
            "战区": current_warzone,
            "分部名称": str(dept) if dept else "",
            "库存金额": safe_float(get_cell_value(sheet, row_idx, 4)),
            "滞销库存": safe_float(get_cell_value(sheet, row_idx, 5)),
            "逾期率": safe_float(get_cell_value(sheet, row_idx, 6)),
        }
        data["data"].append(record)
    
    save_json(data, 'sheet9.json')
    print(f"Sheet9: 共{len(data['data'])}条记录")
    return data

if __name__ == '__main__':
    print("开始处理所有Sheet...")
    print()
    
    process_sheet1()
    process_sheet2()
    process_sheet3()
    process_sheet4()
    process_sheet5()
    process_sheet6()
    process_sheet7()
    process_sheet8()
    process_sheet9()
    
    print()
    print("全部处理完成!")
